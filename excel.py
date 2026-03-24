import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any


class ExcelExporter:

    def __init__(self):
        self.headers = [
            'Ссылка на товар',
            'Артикул',
            'Название',
            'Цена',
            'Описание',
            'Ссылки на изображения',
            'Характеристики',
            'Название селлера',
            'Ссылка на селлера',
            'Размеры',
            'Остатки',
            'Рейтинг',
            'Количество отзывов'
        ]

    def save_catalog(self, products: List[Dict], filename: str):
        # Сохранение каталога в Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Каталог"

        # Заголовки
        for col_idx, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_idx, product in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=product.get('url', ''))
            ws.cell(row=row_idx, column=2, value=product.get('id', ''))
            ws.cell(row=row_idx, column=3, value=product.get('name', ''))
            ws.cell(row=row_idx, column=4, value=product.get('price', ''))
            ws.cell(row=row_idx, column=5, value=product.get('description', ''))
            ws.cell(row=row_idx, column=6, value=product.get('images', ''))
            ws.cell(row=row_idx, column=7, value=product.get('characteristics', ''))
            ws.cell(row=row_idx, column=8, value=product.get('seller_name', ''))
            ws.cell(row=row_idx, column=9, value=product.get('seller_url', ''))
            ws.cell(row=row_idx, column=10, value=product.get('sizes', ''))
            ws.cell(row=row_idx, column=11, value=product.get('total_stock', 0))
            ws.cell(row=row_idx, column=12, value=product.get('rating', 0))
            ws.cell(row=row_idx, column=13, value=product.get('reviews_count', 0))

        # Автоподбор ширины колонок
        for col_idx, header in enumerate(self.headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(header)

            for row_idx in range(2, min(len(products) + 2, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)

    def filter_by_criteria(self, products: List[Dict]) -> List[Dict]:
        """
        Фильтрация товаров по критериям:
        - рейтинг >= 4.5
        - цена <= 10000
        - страна производства Россия
        """
        filtered = []

        for product in products:
            rating = product.get('rating', 0)
            price = product.get('price', float('inf'))

            if rating >= 4.5 and price <= 10000:
                characteristics = product.get('characteristics', '')

                is_russian = False

                if 'страна производства: россия' in characteristics.lower() or \
                        'страна производства:рф' in characteristics.lower() or \
                        'страна производитель: россия' in characteristics.lower():
                    is_russian = True

                if not is_russian:
                    for char in characteristics.split(';'):
                        if 'страна производства' in char.lower() and 'россия' in char.lower():
                            is_russian = True
                            break

                if is_russian:
                    filtered.append(product)

        return filtered